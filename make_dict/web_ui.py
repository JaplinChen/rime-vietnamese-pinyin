import json
import mimetypes
import os
import shutil
import subprocess
import sys
import threading
import traceback
from datetime import datetime
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import openpyxl

from dict_utils import BASE_DIR, ROOT_DIR, clean_tone, remove_tone, to_telex


WORKBOOK_PATH = BASE_DIR / "VietnameseWordList.xlsx"
WEB_DIR = BASE_DIR / "web"
BACKUP_DIR = BASE_DIR / "backups"
WORKBOOK_LOCK = threading.RLock()
ALLOWED_DOWNLOADS = {
    "vn.dict.yaml": ROOT_DIR / "vn.dict.yaml",
    "vn_han.dict.yaml": ROOT_DIR / "vn_han.dict.yaml",
    "make_dict/vn.dict.yaml": BASE_DIR / "vn.dict.yaml",
    "make_dict/vn_han.dict.yaml": BASE_DIR / "vn_han.dict.yaml",
}


def cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def set_cell(sheet, row, column, value):
    cell = sheet.cell(row=row, column=column)
    next_value = value if value != "" else None
    if cell.value == next_value:
        return False
    cell.value = next_value
    return True


def normalize_chinese(value):
    return cell_text(value).replace(", ", "; ").replace("、", "; ")


def has_cjk(text):
    return any("\u4e00" <= char <= "\u9fff" for char in cell_text(text))


def append_import_entry(entries, source, target):
    source = cell_text(source)
    target = normalize_chinese(target)
    if not source:
        return
    if target and source in entries:
        existing_parts = [part.strip() for part in entries[source].split(";") if part.strip()]
        for part in [part.strip() for part in target.split(";") if part.strip()]:
            if part not in existing_parts:
                existing_parts.append(part)
        entries[source] = "; ".join(existing_parts)
        return
    entries.setdefault(source, target)


def append_json_import_entry(entries, source, target):
    source = cell_text(source)
    target = cell_text(target)
    if has_cjk(source) and not target:
        return
    if has_cjk(source) and target and not has_cjk(target):
        append_import_entry(entries, target, source)
    else:
        append_import_entry(entries, source, target)


def backup_workbook(reason):
    if not WORKBOOK_PATH.exists():
        return ""
    if WORKBOOK_PATH.resolve() != (BASE_DIR / "VietnameseWordList.xlsx").resolve():
        return ""
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    safe_reason = "".join(char for char in reason if char.isalnum() or char in {"-", "_"}).strip() or "edit"
    backup_path = BACKUP_DIR / f"VietnameseWordList-{timestamp}-{safe_reason}.xlsx"
    suffix = 1
    while backup_path.exists():
        backup_path = BACKUP_DIR / f"VietnameseWordList-{timestamp}-{safe_reason}-{suffix}.xlsx"
        suffix += 1
    shutil.copy2(WORKBOOK_PATH, backup_path)
    return str(backup_path)


def refresh_derived_columns(sheet):
    last_text = None
    duplicate_count = 0
    dirty = False
    for row_index in range(2, sheet.max_row + 1):
        original_text = cell_text(sheet.cell(row=row_index, column=1).value)
        if not original_text:
            for column in (3, 4, 5, 6):
                dirty |= set_cell(sheet, row_index, column, "")
            continue

        clean_text = clean_tone(original_text)
        no_tone_text = remove_tone(clean_text).strip()
        telex_text = to_telex(clean_text)
        dirty |= set_cell(sheet, row_index, 3, clean_text)
        dirty |= set_cell(sheet, row_index, 4, no_tone_text)
        dirty |= set_cell(sheet, row_index, 5, telex_text)

        if last_text == original_text:
            duplicate_count += 1
            dirty |= set_cell(sheet, row_index, 6, duplicate_count)
        else:
            duplicate_count = 1
            last_text = original_text
            dirty |= set_cell(sheet, row_index, 6, "")

        chinese = sheet.cell(row=row_index, column=2).value
        if chinese is not None:
            dirty |= set_cell(sheet, row_index, 2, normalize_chinese(chinese))
    return dirty


def read_terms():
    with WORKBOOK_LOCK:
        workbook = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
        sheet = workbook.active
        terms = []
        with_chinese = 0
        blank_chinese = 0
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            vietnamese = cell_text(row[0] if len(row) > 0 else "")
            clean = cell_text(row[2] if len(row) > 2 else "")
            if not vietnamese and not clean:
                continue
            chinese = cell_text(row[1] if len(row) > 1 else "")
            if chinese:
                with_chinese += 1
            else:
                blank_chinese += 1
            terms.append(
                {
                    "rowIndex": row_index,
                    "vietnamese": vietnamese,
                    "chinese": chinese,
                    "cleanVietnamese": clean,
                    "pinyin": cell_text(row[3] if len(row) > 3 else ""),
                    "telex": cell_text(row[4] if len(row) > 4 else ""),
                    "duplicate": cell_text(row[5] if len(row) > 5 else ""),
                }
            )
        workbook.close()
    return {
        "terms": terms,
        "stats": {
            "total": len(terms),
            "withChinese": with_chinese,
            "blankChinese": blank_chinese,
            "workbook": str(WORKBOOK_PATH),
        },
    }


def update_term(row_index, payload):
    vietnamese = cell_text(payload.get("vietnamese"))
    chinese = normalize_chinese(payload.get("chinese"))
    if not vietnamese:
        raise ValueError("越南語不可空白")

    with WORKBOOK_LOCK:
        workbook = openpyxl.load_workbook(WORKBOOK_PATH)
        sheet = workbook.active
        if row_index < 2 or row_index > sheet.max_row:
            workbook.close()
            raise ValueError("詞彙列不存在")
        workbook.close()
        backup_path = backup_workbook("update")
        workbook = openpyxl.load_workbook(WORKBOOK_PATH)
        sheet = workbook.active
        set_cell(sheet, row_index, 1, vietnamese)
        set_cell(sheet, row_index, 2, chinese)
        refresh_derived_columns(sheet)
        workbook.save(WORKBOOK_PATH)
        workbook.close()
    return {"ok": True, "rowIndex": row_index, "backup": backup_path}


def add_term(payload):
    vietnamese = cell_text(payload.get("vietnamese"))
    chinese = normalize_chinese(payload.get("chinese"))
    if not vietnamese:
        raise ValueError("越南語不可空白")

    with WORKBOOK_LOCK:
        backup_path = backup_workbook("add")
        workbook = openpyxl.load_workbook(WORKBOOK_PATH)
        sheet = workbook.active
        row_index = sheet.max_row + 1
        set_cell(sheet, row_index, 1, vietnamese)
        set_cell(sheet, row_index, 2, chinese)
        refresh_derived_columns(sheet)
        workbook.save(WORKBOOK_PATH)
        workbook.close()
    return {"ok": True, "rowIndex": row_index, "backup": backup_path}


def clear_chinese(row_index):
    with WORKBOOK_LOCK:
        workbook = openpyxl.load_workbook(WORKBOOK_PATH)
        sheet = workbook.active
        if row_index < 2 or row_index > sheet.max_row:
            workbook.close()
            raise ValueError("詞彙列不存在")
        workbook.close()
        backup_path = backup_workbook("clear")
        workbook = openpyxl.load_workbook(WORKBOOK_PATH)
        sheet = workbook.active
        set_cell(sheet, row_index, 2, "")
        workbook.save(WORKBOOK_PATH)
        workbook.close()
    return {"ok": True, "rowIndex": row_index, "backup": backup_path}


def parse_import(payload):
    text = payload.get("text")
    if not isinstance(text, str):
        raise ValueError("匯入內容不可空白")
    fmt = cell_text(payload.get("format")).lower()
    entries = {}
    skipped = 0
    if fmt == "json" or text.lstrip().startswith("{"):
        decoded = json.loads(text)
        if not isinstance(decoded, dict):
            raise ValueError("JSON 必須是物件")
        if "glossary" in decoded or "langGlossary" in decoded:
            glossary = decoded.get("glossary") or {}
            if not isinstance(glossary, dict):
                raise ValueError("glossary 必須是物件")
            for source, target in glossary.items():
                append_json_import_entry(entries, source, target)
            lang_glossary = decoded.get("langGlossary") or {}
            if isinstance(lang_glossary, dict):
                for glossary_map in lang_glossary.values():
                    if isinstance(glossary_map, dict):
                        for source, target in glossary_map.items():
                            append_json_import_entry(entries, source, target)
        else:
            for source, target in decoded.items():
                append_json_import_entry(entries, source, target)
    else:
        for line in text.splitlines():
            if not line.strip():
                continue
            if "\t" not in line:
                skipped += 1
                continue
            source, target = line.split("\t", 1)
            append_import_entry(entries, source, target)
    return list(entries.items()), skipped


def import_terms(payload):
    entries, skipped = parse_import(payload)
    if not entries:
        raise ValueError("沒有可匯入的詞彙")
    mode = cell_text(payload.get("mode")).lower() or "merge"
    if mode not in {"merge", "replace"}:
        raise ValueError("匯入模式不合法")

    with WORKBOOK_LOCK:
        workbook = openpyxl.load_workbook(WORKBOOK_PATH)
        sheet = workbook.active
        dirty = False
        if mode == "replace":
            for row_index in range(2, sheet.max_row + 1):
                dirty |= set_cell(sheet, row_index, 2, "")
        dirty |= refresh_derived_columns(sheet)
        by_clean = {}
        by_original = {}
        for row_index in range(2, sheet.max_row + 1):
            original = cell_text(sheet.cell(row=row_index, column=1).value)
            clean = cell_text(sheet.cell(row=row_index, column=3).value)
            if original:
                by_original.setdefault(original.lower(), row_index)
            if clean:
                by_clean.setdefault(clean.lower(), row_index)

        imported = 0
        updated = 0
        appended = 0
        for source, target in entries:
            clean = clean_tone(source)
            row_index = by_clean.get(clean.lower()) or by_original.get(source.lower())
            if row_index:
                old_value = cell_text(sheet.cell(row=row_index, column=2).value)
                if mode == "replace" or not old_value:
                    dirty |= set_cell(sheet, row_index, 2, target)
                    updated += int(old_value != target)
            else:
                row_index = sheet.max_row + 1
                dirty |= set_cell(sheet, row_index, 1, source)
                dirty |= set_cell(sheet, row_index, 2, target)
                by_original[source.lower()] = row_index
                by_clean[clean.lower()] = row_index
                appended += 1
            imported += 1

        dirty |= refresh_derived_columns(sheet)
        backup_path = backup_workbook(f"import-{mode}") if dirty else ""
        if dirty:
            workbook.save(WORKBOOK_PATH)
        workbook.close()
    return {
        "ok": True,
        "imported": imported,
        "updated": updated,
        "appended": appended,
        "skipped": skipped,
        "mode": mode,
        "backup": backup_path,
    }


def export_bundle():
    data = read_terms()
    entries = [
        (term["cleanVietnamese"] or term["vietnamese"], term["chinese"])
        for term in data["terms"]
        if term["chinese"]
    ]
    glossary = {source: target for source, target in entries}
    tsv = "\n".join(f"{source}\t{target}" for source, target in entries)
    return {"json": json.dumps(glossary, ensure_ascii=False, indent=2), "tsv": tsv, "count": len(entries)}


def run_command(args):
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    result = subprocess.run(
        args,
        cwd=ROOT_DIR,
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
    )
    return {"ok": result.returncode == 0, "returnCode": result.returncode, "output": result.stdout}


def generate_yaml(payload):
    args = [sys.executable, str(BASE_DIR / "excel_to_vndict.py")]
    if payload.get("syncRoot", True):
        args.append("--sync-root")
    with WORKBOOK_LOCK:
        return run_command(args)


def verify_project():
    return run_command([sys.executable, str(BASE_DIR / "verify.py")])


class WebUiHandler(SimpleHTTPRequestHandler):
    server_version = "RimeVietnameseWebUI/0.1"

    def log_message(self, format, *args):
        print(f"{self.address_string()} - {format % args}")

    def send_json(self, body, status=HTTPStatus.OK):
        raw = json.dumps(body, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def send_error_json(self, exc, status=HTTPStatus.BAD_REQUEST):
        self.send_json({"ok": False, "error": str(exc)}, status)

    def read_json(self):
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length).decode("utf-8") if length else "{}"
        return json.loads(raw or "{}")

    def do_GET(self):
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/terms":
                self.send_json(read_terms())
                return
            if parsed.path == "/api/export":
                self.send_json(export_bundle())
                return
            if parsed.path == "/api/download":
                query = parse_qs(parsed.query)
                key = query.get("file", [""])[0]
                path = ALLOWED_DOWNLOADS.get(key)
                if path is None or not path.exists():
                    self.send_error_json("檔案不存在", HTTPStatus.NOT_FOUND)
                    return
                raw = path.read_bytes()
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "text/yaml; charset=utf-8")
                self.send_header("Content-Disposition", f'attachment; filename="{path.name}"')
                self.send_header("Content-Length", str(len(raw)))
                self.end_headers()
                self.wfile.write(raw)
                return
            self.serve_static(parsed.path)
        except Exception as exc:
            traceback.print_exc()
            self.send_error_json(exc, HTTPStatus.INTERNAL_SERVER_ERROR)

    def do_POST(self):
        parsed = urlparse(self.path)
        try:
            payload = self.read_json()
            if parsed.path == "/api/terms":
                self.send_json(add_term(payload))
                return
            if parsed.path == "/api/import":
                self.send_json(import_terms(payload))
                return
            if parsed.path == "/api/generate":
                self.send_json(generate_yaml(payload))
                return
            if parsed.path == "/api/verify":
                self.send_json(verify_project())
                return
            self.send_error_json("找不到 API", HTTPStatus.NOT_FOUND)
        except Exception as exc:
            traceback.print_exc()
            self.send_error_json(exc)

    def do_PUT(self):
        parsed = urlparse(self.path)
        try:
            parts = parsed.path.strip("/").split("/")
            if len(parts) == 3 and parts[:2] == ["api", "terms"]:
                self.send_json(update_term(int(parts[2]), self.read_json()))
                return
            self.send_error_json("找不到 API", HTTPStatus.NOT_FOUND)
        except Exception as exc:
            traceback.print_exc()
            self.send_error_json(exc)

    def do_DELETE(self):
        parsed = urlparse(self.path)
        try:
            parts = parsed.path.strip("/").split("/")
            if len(parts) == 3 and parts[:2] == ["api", "terms"]:
                self.send_json(clear_chinese(int(parts[2])))
                return
            self.send_error_json("找不到 API", HTTPStatus.NOT_FOUND)
        except Exception as exc:
            traceback.print_exc()
            self.send_error_json(exc)

    def serve_static(self, request_path):
        if request_path in {"", "/"}:
            request_path = "/index.html"
        relative = request_path.lstrip("/")
        root = WEB_DIR.resolve()
        path = (root / relative).resolve()
        try:
            path.relative_to(root)
        except ValueError:
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        if not path.exists() or path.is_dir():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        raw = path.read_bytes()
        mime_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        if path.suffix in {".html", ".css", ".js"}:
            mime_type += "; charset=utf-8"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", mime_type)
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)


def parse_args():
    import argparse

    parser = argparse.ArgumentParser(description="啟動 Rime 越南語詞彙表 Web UI。")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8765)
    return parser.parse_args()


def main():
    args = parse_args()
    server = ThreadingHTTPServer((args.host, args.port), WebUiHandler)
    print(f"詞彙表 Web UI：http://{args.host}:{args.port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
