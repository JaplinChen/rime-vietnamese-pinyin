import py_compile
import os
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl


BASE_DIR = Path(__file__).resolve().parent
REPO_ROOT = BASE_DIR.parent

PYTHON_FILES = (
    BASE_DIR / "dict_utils.py",
    BASE_DIR / "excel_to_vndict.py",
    BASE_DIR / "vn_to_pinyin.py",
    BASE_DIR / "vn_to_telex.py",
    BASE_DIR / "validate_dicts.py",
    BASE_DIR / "web_ui.py",
    BASE_DIR / "import_typetwo.py",
)


def configure_output_encoding():
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8")


def compile_scripts():
    print("檢查 Python 腳本語法...")
    for path in PYTHON_FILES:
        py_compile.compile(path, doraise=True)


def run_command(args, expect_success=True):
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"

    result = subprocess.run(
        args,
        cwd=REPO_ROOT,
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
    )

    if expect_success and result.returncode != 0:
        print(result.stdout, end="")
        raise SystemExit(result.returncode)

    if not expect_success and result.returncode == 0:
        command = " ".join(str(arg) for arg in args)
        raise SystemExit(f"命令應該失敗但成功了：{command}")

    if result.stdout and expect_success:
        print(result.stdout, end="")

    return result


def validate_dictionaries():
    print("驗證字典格式與根目錄同步狀態...")
    run_command([sys.executable, str(BASE_DIR / "validate_dicts.py")])


def validate_missing_word_list_guard():
    print("確認輔助腳本在缺少 WordList.txt 時不會覆寫輸出...")
    run_command([sys.executable, str(BASE_DIR / "vn_to_pinyin.py")], expect_success=False)
    run_command([sys.executable, str(BASE_DIR / "vn_to_telex.py")], expect_success=False)


def validate_web_ui_import_parser():
    print("驗證 Web UI 詞彙匯入解析...")
    sys.path.insert(0, str(BASE_DIR))
    from web_ui import parse_import

    typetwo_entries, skipped = parse_import(
        {
            "format": "json",
            "text": '{"上傳": "Tải lên", "上傳檔案": "Tải lên"}',
        }
    )
    if skipped != 0 or typetwo_entries != [("Tải lên", "上傳; 上傳檔案")]:
        raise SystemExit(f"TypeTwo JSON 匯入解析錯誤：{typetwo_entries}, skipped={skipped}")

    exported_entries, skipped = parse_import(
        {
            "format": "json",
            "text": '{"Tải lên": "上傳"}',
        }
    )
    if skipped != 0 or exported_entries != [("Tải lên", "上傳")]:
        raise SystemExit(f"匯出 JSON 再匯入解析錯誤：{exported_entries}, skipped={skipped}")

    tsv_entries, skipped = parse_import(
        {
            "format": "tsv",
            "text": "Tải lên\t上傳\n無效行",
        }
    )
    if skipped != 1 or tsv_entries != [("Tải lên", "上傳")]:
        raise SystemExit(f"TSV 匯入解析錯誤：{tsv_entries}, skipped={skipped}")


def validate_web_ui_import_modes():
    print("驗證 Web UI 詞彙匯入模式...")
    sys.path.insert(0, str(BASE_DIR))
    import web_ui

    original_workbook_path = web_ui.WORKBOOK_PATH
    original_backup_dir = web_ui.BACKUP_DIR
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir) / "VietnameseWordList.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Vietnamese", "Chinese", "Clean Vietnamese", "Pinyin", "Telex", "Duplicate"])
        sheet.append(["Tải lên", "既有翻譯"])
        sheet.append(["Tải xuống", None])
        workbook.save(temp_path)
        workbook.close()

        try:
            web_ui.WORKBOOK_PATH = temp_path
            web_ui.BACKUP_DIR = Path(temp_dir) / "backups"

            merge_result = web_ui.import_terms(
                {
                    "format": "json",
                    "mode": "merge",
                    "text": '{"上傳": "Tải lên", "下載": "Tải xuống", "新增": "Tạo mới"}',
                }
            )
            workbook = openpyxl.load_workbook(temp_path)
            sheet = workbook.active
            merge_values = [sheet.cell(row=row, column=2).value for row in range(2, 5)]
            workbook.close()
            if merge_result["updated"] != 1 or merge_result["appended"] != 1:
                raise SystemExit(f"merge 匯入計數錯誤：{merge_result}")
            if merge_values != ["既有翻譯", "下載", "新增"]:
                raise SystemExit(f"merge 匯入覆寫或補值錯誤：{merge_values}")

            replace_result = web_ui.import_terms(
                {
                    "format": "json",
                    "mode": "replace",
                    "text": '{"上傳": "Tải lên"}',
                }
            )
            workbook = openpyxl.load_workbook(temp_path)
            sheet = workbook.active
            replace_values = [sheet.cell(row=row, column=2).value for row in range(2, 5)]
            workbook.close()
            if replace_result["updated"] != 1:
                raise SystemExit(f"replace 匯入計數錯誤：{replace_result}")
            if replace_values != ["上傳", None, None]:
                raise SystemExit(f"replace 匯入清空或套用錯誤：{replace_values}")
        finally:
            web_ui.WORKBOOK_PATH = original_workbook_path
            web_ui.BACKUP_DIR = original_backup_dir


def main():
    configure_output_encoding()
    compile_scripts()
    validate_web_ui_import_parser()
    validate_web_ui_import_modes()
    validate_dictionaries()
    validate_missing_word_list_guard()
    print("驗證完成")


if __name__ == "__main__":
    main()
