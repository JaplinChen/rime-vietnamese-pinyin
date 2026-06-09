import argparse
from pathlib import Path

import openpyxl

from dict_utils import BASE_DIR, ROOT_DIR, clean_tone, copy_text_atomic, remove_tone, to_telex, write_text_atomic


WORKBOOK_PATH = BASE_DIR / "VietnameseWordList.xlsx"
VN_DICT_PATH = BASE_DIR / "vn.dict.yaml"
VN_HAN_DICT_PATH = BASE_DIR / "vn_han.dict.yaml"
ROOT_VN_DICT_PATH = ROOT_DIR / "vn.dict.yaml"
ROOT_VN_HAN_DICT_PATH = ROOT_DIR / "vn_han.dict.yaml"

VN_HEADER = """# Rime dictionary
# encoding: utf-8
---
name: vn
version: "2020.05.29"
sort: original
use_preset_vocabulary: false
max_phrase_length: 10
min_phrase_weight: 100
...

"""

VN_HAN_HEADER = """# Rime dictionary
# encoding: utf-8
---
name: vn_han
version: "2020.05.29"
sort: original
use_preset_vocabulary: false
...

"""


def set_cell_value(sheet, row, column, value):
    cell = sheet.cell(row=row, column=column)
    if cell.value == value:
        return False

    cell.value = value
    return True


def parse_args():
    parser = argparse.ArgumentParser(description="從 Excel 重新產生 Rime 越南語字典。")
    parser.add_argument(
        "--sync-root",
        action="store_true",
        help="產生後同步更新專案根目錄的 vn.dict.yaml 與 vn_han.dict.yaml。",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    workbook = openpyxl.load_workbook(WORKBOOK_PATH)
    sheet = workbook.active

    vn_lines = [VN_HEADER]
    vn_han_lines = [VN_HAN_HEADER]
    last_text = None
    duplicate_count = 0
    vietnamese_entry_count = 0
    chinese_entry_count = 0
    workbook_dirty = False

    for row_index, (vietnamese_cell,) in enumerate(
        sheet.iter_rows(min_col=1, max_col=1),
        start=1,
    ):
        if row_index == 1 or vietnamese_cell.value is None:
            continue

        original_text = str(vietnamese_cell.value).strip()
        if not original_text:
            continue

        clean_text = clean_tone(original_text)
        no_tone_text = remove_tone(clean_text).strip()
        telex_text = to_telex(clean_text)

        workbook_dirty |= set_cell_value(sheet, row_index, 3, clean_text)
        workbook_dirty |= set_cell_value(sheet, row_index, 4, no_tone_text)
        workbook_dirty |= set_cell_value(sheet, row_index, 5, telex_text)

        if last_text == original_text:
            duplicate_count += 1
            workbook_dirty |= set_cell_value(sheet, row_index, 6, duplicate_count)
        else:
            duplicate_count = 1
            last_text = original_text
            workbook_dirty |= set_cell_value(sheet, row_index, 6, None)

        vn_lines.append(f"{clean_text} \t{telex_text}\t50000\n")
        vn_lines.append(f"{clean_text} \t{no_tone_text}\t40000\n")
        vietnamese_entry_count += 2

        chinese_cell = sheet.cell(row=row_index, column=2)
        if chinese_cell.value is None:
            continue

        chinese_text = str(chinese_cell.value).replace(", ", "; ").replace("、", "; ").strip()
        workbook_dirty |= set_cell_value(sheet, row_index, 2, chinese_text)

        vn_han_lines.append(f"{chinese_text}\t{clean_text}\t30000\n")
        vn_han_lines.append(f"{chinese_text}\t{telex_text}\t20000\n")
        vn_han_lines.append(f"{chinese_text}\t{no_tone_text}\t10000\n")
        chinese_entry_count += 3

    if workbook_dirty:
        workbook.save(WORKBOOK_PATH)
    write_text_atomic(VN_DICT_PATH, "".join(vn_lines))
    write_text_atomic(VN_HAN_DICT_PATH, "".join(vn_han_lines))

    if args.sync_root:
        copy_text_atomic(VN_DICT_PATH, ROOT_VN_DICT_PATH)
        copy_text_atomic(VN_HAN_DICT_PATH, ROOT_VN_HAN_DICT_PATH)

    print(f"vn.dict.yaml：{vietnamese_entry_count} 筆")
    print(f"vn_han.dict.yaml：{chinese_entry_count} 筆")
    if workbook_dirty:
        print("已更新 VietnameseWordList.xlsx 衍生欄位")
    else:
        print("VietnameseWordList.xlsx 衍生欄位無變更")
    if args.sync_root:
        print("已同步根目錄發布字典")


if __name__ == "__main__":
    main()
